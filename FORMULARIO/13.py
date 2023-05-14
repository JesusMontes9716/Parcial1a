import openpyxl as xl
from openpyxl import *

from tkinter import messagebox
from tkinter import*
from tkinter import ttk

wb = load_workbook('formulario.xlsx')
sheet = wb.active

def excel():
     
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "NOMBRE"
    sheet.cell(row=1, column=2).value = "AP. PATERNO"
    sheet.cell(row=1, column=3).value = "AP. MATERNO"
    sheet.cell(row=1, column=4).value = "CORREO"
    sheet.cell(row=1, column=5).value = "MOVIL"
    sheet.cell(row=1, column=6).value = "AFICIONES"
    sheet.cell(row=1, column=7).value = "ESTADO"
    sheet.cell(row=1, column=8).value = "ESTATUS"

def insert():
     
    # if user not fill any entry
    # then print "empty input"
    if (nombreEntry.get() == "" and
        aPaternoEntry.get() == "" and
        aMaternoEntry.get() == "" and
        correoEntry.get() == "" and
        movilEntry.get() == "" 
        ):
             
        print("empty input")
 
    else:
 
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
 
        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = nombreEntry.get()
        sheet.cell(row=current_row + 1, column=2).value = aPaternoEntry.get()
        sheet.cell(row=current_row + 1, column=3).value = aMaternoEntry.get()
        sheet.cell(row=current_row + 1, column=4).value = correoEntry.get()
        sheet.cell(row=current_row + 1, column=5).value = movilEntry.get()
        
 
        # save the file
        wb.save('formulario.xlsx')
 
        # set focus on the name_field box
        nombreEntry.focus_set()
 
        
 
 
raiz = Tk()

estado = StringVar()
mainFrame = ttk.Frame(raiz, padding="10 10 10 10")#izquierda arriba derecha abajo
mainFrame.grid(column=0, row=0)

frame1 = ttk.Frame(mainFrame, padding="10 10 10 10", borderwidth=3, relief="raised")
frame1.grid(column=0, row=0)

frame2 = ttk.Frame(mainFrame, padding="10 10 10 10", borderwidth=3, relief="raised")
frame2.grid(column=0, row=1)
frame3 = ttk.Frame(mainFrame, padding="10 10 10 10")
frame3.grid(column=1, row=0)
        
nombreEntry = ttk.Entry(frame1, width=30)
nombreEntry.grid(column=1, row=0, pady=12, sticky=(W, E))

aPaternoEntry = ttk.Entry(frame1, width=30)
aPaternoEntry.grid(column=1, row=2, pady=12, sticky=(W, E))

aMaternoEntry = ttk.Entry(frame1, width=30)
aMaternoEntry.grid(column=1, row=3, pady=12, sticky=(W, E))

correoEntry = ttk.Entry(frame1, width=30)
correoEntry.grid(column=1, row=4, pady=12, sticky=(W, E))

movilEntry = ttk.Entry(frame1, width=30)
movilEntry.grid(column=1, row=5, pady=12, sticky=(W, E))
        
ttk.Label(frame1, text="NOMBRE:  ").grid(column=0, row=0, pady=12, sticky=(E))
ttk.Label(frame1, text="AP. PATERNO:  ").grid(column=0, row=2, pady=12, sticky=(E))
ttk.Label(frame1, text="AP. MATERNO:  ").grid(column=0, row=3, pady=12, sticky=(E))
ttk.Label(frame1, text="CORREO:  ").grid(column=0, row=4, pady=12, sticky=(E))
ttk.Label(frame1, text="MOVIL:  ").grid(column=0, row=5, pady=12, sticky=(E))

ttk.Label(frame2, text="Aficiones:  ").grid(column=0, row=0, pady=12, sticky=(E))
leer = ttk.Checkbutton(frame2, text='LEER')
leer.grid(column=0, row=1, sticky=(W))
musica = ttk.Checkbutton(frame2, text='MUSICA')
musica.grid(column=1, row=1, sticky=(W))
videojuegos = ttk.Checkbutton(frame2, text='VIDEOJUEGOS')
videojuegos.grid(column=2, row=1, sticky=(W))

        
comboEstados = ttk.Combobox(mainFrame, textvariable=estado)
comboEstados.grid(column=1, row=1, pady=50)
comboEstados['values']=('Aguascalientes','Baja California','Baja California Sur','Campeche',
'Chiapas','Chihuahua','Ciudad de México','Coahuila','Colima','Durango','Guanajuato','Guerrero',
'Hidalgo','Jalisco','México','Michoacán','Morelos','Nayarit','Nuevo León','Oaxaca','Puebla',
'Querétaro','Quintana Roo','San Luis Potosí','Sinaloa','Sonora','Tabasco','Tamaulipas',
'Tlaxcala','Veracruz','Yucatán','Zacatecas')
        
ttk.Button(mainFrame ,text="GUARDAR",).grid(column=0, row=5, pady=12, sticky=(W))
ttk.Button(mainFrame, text="CANCELAR",).grid(column=0, row=5, pady=12, sticky=(N))

estudiante = ttk.Radiobutton(frame3, text='ESTUDIANTE')
estudiante.grid(column=0, row=0, sticky=(W))
empleado = ttk.Radiobutton(frame3, text='EMPLEADO')
empleado.grid(column=0, row=1, sticky=(W))
desempleado = ttk.Radiobutton(frame3, text='DESEMPLEADO')
desempleado.grid(column=0, row=2, sticky=(W))

nombreEntry.focus()

raiz.mainloop()